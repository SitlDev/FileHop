from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from io import BytesIO, StringIO
from datetime import datetime
from typing import List, Dict, Optional
import csv
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ReportGenerator:
    """
    CMS Five-Star Quality Rating Report Generator
    Formats reports to match official CMS Nursing Home Compare format
    """
    def __init__(self, title: str = "CMS Five-Star Quality Rating Report"):
        self.title = title
        self.pagesize = letter
        self.cms_blue = colors.HexColor('#003366')
        self.cms_light_gray = colors.HexColor('#f5f5f5')
        
    def generate_facility_report(
        self,
        facility_name: str,
        facility_data: Dict,
        ratings_data: List[Dict],
        inspections_data: List[Dict]
    ) -> BytesIO:
        """Generate CMS-compliant facility report with Five-Star ratings"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=self.pagesize,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.6*inch,
            bottomMargin=0.5*inch,
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # CMS-style header
        cms_header = ParagraphStyle(
            'CMSHeader',
            parent=styles['Heading1'],
            fontSize=10,
            textColor=colors.black,
            spaceAfter=3,
            fontName='Helvetica-Bold'
        )
        
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=self.cms_blue,
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=self.cms_blue,
            spaceAfter=10,
            spaceBefore=12,
            fontName='Helvetica-Bold',
            borderPadding=5,
            backColor=self.cms_light_gray
        )
        
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=9,
            leading=11
        )
        
        # CMS Header
        story.append(Paragraph("Centers for Medicare & Medicaid Services", cms_header))
        story.append(Paragraph("FIVE-STAR QUALITY RATING SYSTEM - NURSING HOME COMPARE", cms_header))
        story.append(Spacer(1, 0.1*inch))
        
        # Report Title
        story.append(Paragraph("DETAILED FACILITY REPORT", title_style))
        story.append(Spacer(1, 0.15*inch))
        
        # Facility Identification Section
        story.append(Paragraph("FACILITY IDENTIFICATION", heading_style))
        
        cms_provider_id = facility_data.get('cms_provider_id', 'N/A')
        report_date = datetime.now().strftime('%B %d, %Y')
        
        # Extract address components
        address_dict = facility_data.get('address', {})
        if isinstance(address_dict, dict):
            street = address_dict.get('street', 'N/A')
            city = address_dict.get('city', 'N/A')
            state = address_dict.get('state', 'N/A')
            zip_code = address_dict.get('zip', 'N/A')
        else:
            street = str(address_dict) if address_dict else 'N/A'
            city = facility_data.get('city', 'N/A')
            state = facility_data.get('state', 'N/A')
            zip_code = facility_data.get('zip_code', 'N/A')
        
        facility_info = [
            ['Facility Name:', facility_name],
            ['CMS Provider ID:', cms_provider_id],
            ['Address:', street],
            ['City, State, ZIP:', f"{city}, {state} {zip_code}"],
            ['Licensed Beds:', str(facility_data.get('bed_count', 'N/A'))],
            ['Report Date:', report_date],
        ]
        
        facility_table = Table(facility_info, colWidths=[1.8*inch, 4.2*inch])
        facility_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LINEABOVE', (0, 0), (-1, -1), 0.5, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 0.5, colors.black),
        ]))
        story.append(facility_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Overall Star Rating - Most Prominent Section
        if ratings_data:
            latest_rating = ratings_data[0]
            overall_stars = latest_rating.get('overall_rating', 0)
            
            story.append(Paragraph("OVERALL RATING", heading_style))
            
            # Create star rating display using the unified method
            star_display = self._star_string(overall_stars)
            
            overall_para = ParagraphStyle(
                'OverallRating',
                parent=styles['Normal'],
                fontSize=36,
                textColor=self.cms_blue,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold',
                spaceAfter=6,
                spaceBefore=6,
                leading=40
            )
            
            stars_numeric = ParagraphStyle(
                'OverallNumeric',
                parent=styles['Normal'],
                fontSize=16,
                textColor=self.cms_blue,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold',
                spaceAfter=12,
                spaceBefore=4
            )
            
            story.append(Paragraph(star_display, overall_para))
            story.append(Spacer(1, 0.08*inch))
            story.append(Paragraph(f"{int(overall_stars)}/5 Stars", stars_numeric))
            story.append(Spacer(1, 0.1*inch))
            
            # Four-Domain Star Ratings
            story.append(Paragraph("FOUR-DOMAIN STAR RATINGS", heading_style))
            
            domains = [
                ['Domain', 'Rating', 'Star Display', 'Effective Date'],
                [
                    'Health Inspections',
                    f"{latest_rating.get('health_inspection_rating', 'N/A')} Stars",
                    self._star_string(latest_rating.get('health_inspection_rating', 0)),
                    str(latest_rating.get('effective_date', 'N/A'))[:10]
                ],
                [
                    'Staffing',
                    f"{latest_rating.get('staffing_rating', 'N/A')} Stars",
                    self._star_string(latest_rating.get('staffing_rating', 0)),
                    str(latest_rating.get('effective_date', 'N/A'))[:10]
                ],
                [
                    'Quality Measures',
                    f"{latest_rating.get('qm_rating', 'N/A')} Stars",
                    self._star_string(latest_rating.get('qm_rating', 0)),
                    str(latest_rating.get('effective_date', 'N/A'))[:10]
                ],
            ]
            
            domains_table = Table(domains, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1.6*inch])
            domains_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.cms_blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, self.cms_light_gray]),
            ]))
            story.append(domains_table)
            story.append(Spacer(1, 0.2*inch))
        
        # Health Inspections Domain Section
        if inspections_data:
            story.append(PageBreak())
            story.append(Paragraph("HEALTH INSPECTIONS DOMAIN", heading_style))
            
            # Summary statistics
            total_inspections = len(inspections_data)
            total_deficiencies = sum(len(insp.get('deficiencies', [])) for insp in inspections_data)
            
            summary_data = [
                ['Total Inspections Reviewed:', str(total_inspections)],
                ['Total Deficiencies Found:', str(total_deficiencies)],
                ['Average Deficiencies per Inspection:', f"{total_deficiencies/max(total_inspections, 1):.2f}"],
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (1, 0), (1, -1), self.cms_light_gray),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.15*inch))
            
            # Recent Inspections Table
            story.append(Paragraph("Recent Inspections (Last 3 Years)", ParagraphStyle('SubHeading', parent=styles['Heading3'], fontSize=10, fontName='Helvetica-Bold')))
            
            inspection_rows = [['Inspection Date', 'Type', 'Deficiencies', 'Status']]
            for inspection in sorted(inspections_data, key=lambda x: x.get('inspection_date', ''), reverse=True)[:20]:
                insp_date = str(inspection.get('inspection_date', 'N/A'))[:10]
                insp_type = inspection.get('inspection_type', 'N/A')
                deficiency_count = len(inspection.get('deficiencies', []))
                status = inspection.get('status', 'N/A')
                
                inspection_rows.append([
                    insp_date,
                    insp_type,
                    str(deficiency_count),
                    status
                ])
            
            inspection_table = Table(inspection_rows, colWidths=[1.4*inch, 1.6*inch, 1.3*inch, 1.7*inch])
            inspection_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.cms_blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, self.cms_light_gray]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(inspection_table)
        
        # Methodology Footer
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("___" * 35, styles['Normal']))
        
        methodology_style = ParagraphStyle(
            'Methodology',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.grey,
            alignment=TA_JUSTIFY,
            leading=9
        )
        
        methodology_text = "The Five-Star Quality Rating System provides consumers, their families, and caregivers with information to help compare nursing homes. The system uses data about the nursing home's staffing, health and safety inspections, quality of care indicators, and resident and family satisfaction. Nursing home ratings are updated on a quarterly basis. For the most current information, visit www.medicare.gov/nursinghomecompare."
        
        story.append(Paragraph(methodology_text, methodology_style))
        story.append(Spacer(1, 0.1*inch))
        
        footer_date = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        story.append(Paragraph(f"Report Generated: {footer_date}", footer_style))
        story.append(Paragraph("Source: Centers for Medicare & Medicaid Services (CMS)", footer_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_ratings_trend_report(
        self,
        facility_name: str,
        ratings_history: List[Dict]
    ) -> BytesIO:
        """Generate CMS-formatted ratings trend report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=self.pagesize,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.6*inch,
            bottomMargin=0.5*inch,
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        cms_header = ParagraphStyle(
            'CMSHeader',
            parent=styles['Heading1'],
            fontSize=10,
            textColor=colors.black,
            spaceAfter=3,
            fontName='Helvetica-Bold'
        )
        
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=self.cms_blue,
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=self.cms_blue,
            spaceAfter=10,
            spaceBefore=12,
            fontName='Helvetica-Bold',
            borderPadding=5,
            backColor=self.cms_light_gray
        )
        
        # CMS Header
        story.append(Paragraph("Centers for Medicare & Medicaid Services", cms_header))
        story.append(Paragraph("FIVE-STAR QUALITY RATING SYSTEM - NURSING HOME COMPARE", cms_header))
        story.append(Spacer(1, 0.1*inch))
        
        # Title
        story.append(Paragraph("STAR RATING TREND ANALYSIS", title_style))
        story.append(Paragraph(facility_name, styles['Heading2']))
        story.append(Spacer(1, 0.15*inch))
        
        # Facility Information
        story.append(Paragraph("RATING HISTORY", heading_style))
        
        if ratings_history:
            sorted_ratings = sorted(ratings_history, key=lambda x: x.get('effective_date', ''), reverse=True)
            
            rows = [['Effective Date', 'Overall', 'Health Insp', 'Staffing', 'QM', 'Trend']]
            
            for i, rating in enumerate(sorted_ratings[:24]):  # Last 24 ratings (2 years)
                date_str = str(rating.get('effective_date', 'N/A'))[:10]
                overall = str(int(rating.get('overall_rating', 0)))
                health = str(int(rating.get('health_inspection_rating', 0)))
                staffing = str(int(rating.get('staffing_rating', 0)))
                qm = str(int(rating.get('qm_rating', 0)))
                
                # Calculate trend
                if i < len(sorted_ratings) - 1:
                    current_overall = rating.get('overall_rating', 0)
                    previous_overall = sorted_ratings[i + 1].get('overall_rating', 0)
                    if current_overall > previous_overall:
                        trend = 'Improved'
                    elif current_overall < previous_overall:
                        trend = 'Declined'
                    else:
                        trend = 'Stable'
                else:
                    trend = 'Current'
                
                rows.append([date_str, overall, health, staffing, qm, trend])
            
            table = Table(rows, colWidths=[1.3*inch, 0.8*inch, 1*inch, 0.9*inch, 0.8*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.cms_blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, self.cms_light_gray]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(table)
            
            # Summary
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph("SUMMARY", heading_style))
            
            latest = sorted_ratings[0]
            oldest = sorted_ratings[-1]
            overall_change = latest.get('overall_rating', 0) - oldest.get('overall_rating', 0)
            
            summary_text = f"<b>Period Reviewed:</b> {len(sorted_ratings)} rating periods<br/>" \
                          f"<b>Starting Date:</b> {str(oldest.get('effective_date', 'N/A'))[:10]}<br/>" \
                          f"<b>Current Date:</b> {str(latest.get('effective_date', 'N/A'))[:10]}<br/>" \
                          f"<b>Overall Rating Change:</b> {int(overall_change):+d} stars"
            
            story.append(Paragraph(summary_text, styles['Normal']))
        
        # Footer
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("___" * 35, styles['Normal']))
        
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        footer_date = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        story.append(Paragraph(f"Report Generated: {footer_date}", footer_style))
        story.append(Paragraph("Source: Centers for Medicare & Medicaid Services (CMS)", footer_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def _star_string(self, rating: int) -> str:
        """Convert numeric rating to star string - display only filled stars"""
        rating = int(rating) if isinstance(rating, (int, float)) else 0
        # Ensure rating is in valid 0-5 range
        rating = max(0, min(5, rating))
        
        # Use filled star only
        filled_star = "★"  # U+2605 BLACK STAR - filled star
        
        # Return only filled stars based on rating (no hollow stars)
        return filled_star * rating
    
    def generate_ratings_trend_report(
        self,
        facility_name: str,
        ratings_history: List[Dict]
    ) -> BytesIO:
        """Generate ratings trend analysis report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=self.pagesize,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch,
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1e293b'),
            spaceAfter=6,
            alignment=TA_CENTER,
        )
        
        # Title
        story.append(Paragraph("Star Ratings Trend Analysis", title_style))
        story.append(Paragraph(f"{facility_name}", styles['Heading2']))
        story.append(Paragraph(f"Report Period: {len(ratings_history)} ratings reviewed", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Ratings table
        if ratings_history:
            rows = [['Date', 'Overall', 'Health Insp', 'Staffing', 'QM/Report', 'Trend']]
            for rating in sorted(ratings_history, key=lambda x: x.get('effective_date', ''), reverse=True)[:12]:
                rows.append([
                    rating.get('effective_date', 'N/A')[:10],
                    str(rating.get('overall_rating', 'N/A')),
                    str(rating.get('health_inspection_rating', 'N/A')),
                    str(rating.get('staffing_rating', 'N/A')),
                    str(rating.get('qm_rating', 'N/A')),
                    '📈 Improving' if rating.get('trend') == 'up' else '📉 Declining' if rating.get('trend') == 'down' else '➡️ Stable'
                ])
            
            table = Table(rows, colWidths=[1.2*inch, 0.9*inch, 1*inch, 0.9*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
            ]))
            story.append(table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_staffing_report(self, facility_name: str, staffing_data: List[Dict], benchmarks: Dict, include_comparative: bool = False) -> BytesIO:
        """Generate CMS-compliant staffing domain report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=self.pagesize, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.6*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        heading_style = ParagraphStyle('SectionHeading', parent=styles['Heading2'], fontSize=12, textColor=self.cms_blue, spaceAfter=10, spaceBefore=12, fontName='Helvetica-Bold', backColor=self.cms_light_gray, borderPadding=5)
        
        # CMS Header
        story.append(Paragraph("Centers for Medicare & Medicaid Services", ParagraphStyle('CMSHeader', parent=styles['Heading1'], fontSize=10, textColor=colors.black, spaceAfter=3, fontName='Helvetica-Bold')))
        story.append(Paragraph("FIVE-STAR QUALITY RATING SYSTEM - NURSING HOME COMPARE", ParagraphStyle('CMSHeader', parent=styles['Heading1'], fontSize=10, textColor=colors.black, spaceAfter=3, fontName='Helvetica-Bold')))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("STAFFING DOMAIN REPORT", ParagraphStyle('ReportTitle', parent=styles['Heading1'], fontSize=16, textColor=self.cms_blue, spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')))
        story.append(Paragraph(facility_name, styles['Heading2']))
        story.append(Spacer(1, 0.15*inch))
        
        # Staffing Levels
        story.append(Paragraph("STAFFING LEVELS", heading_style))
        if staffing_data:
            current = staffing_data[0]
            staffing_info = [
                ['Registered Nurses (RN)', str(current.get('rn_count', 'N/A'))],
                ['Licensed Practical Nurses (LPN)', str(current.get('lpn_count', 'N/A'))],
                ['Certified Nursing Assistants (CNA)', str(current.get('cna_count', 'N/A'))],
                ['RN Hours per 100 Resident Days', f"{current.get('rn_hours_per_100_bed_days', 'N/A'):.2f}"],
                ['Total Staff Hours per 100 Resident Days', f"{current.get('total_hours_per_100_bed_days', 'N/A'):.2f}"],
            ]
            staffing_table = Table(staffing_info, colWidths=[3.5*inch, 1.5*inch])
            staffing_table.setStyle(TableStyle([('ALIGN', (0, 0), (0, -1), 'LEFT'), ('ALIGN', (1, 0), (1, -1), 'CENTER'), ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, self.cms_light_gray])]))
            story.append(staffing_table)
            story.append(Spacer(1, 0.15*inch))
        
        # Benchmarking
        if include_comparative and benchmarks:
            story.append(Paragraph("BENCHMARKING", heading_style))
            if staffing_data and staffing_data[0].get('rn_hours_per_100_bed_days') and benchmarks.get('national'):
                current = staffing_data[0]
                bench_data = [
                    ['Measure', 'Your Facility', 'State Median', 'National Median'],
                    ['RN Hours per 100 Bed Days', f"{current.get('rn_hours_per_100_bed_days', 'N/A'):.2f}", 
                     f"{benchmarks.get('state', [{}])[0].get('rn_hours_median', 'N/A'):.2f}" if benchmarks.get('state') else 'N/A',
                     f"{benchmarks.get('national', {}).get('rn_hours_median', 'N/A'):.2f}"],
                ]
                bench_table = Table(bench_data, colWidths=[2*inch, 1.3*inch, 1.3*inch, 1.3*inch])
                bench_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), self.cms_blue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, self.cms_light_gray])]))
                story.append(bench_table)
        
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("___" * 35, styles['Normal']))
        methodology_style = ParagraphStyle('Methodology', parent=styles['Normal'], fontSize=7, textColor=colors.grey, alignment=TA_JUSTIFY, leading=9)
        story.append(Paragraph("The Five-Star Quality Rating System provides consumers, their families, and caregivers with information to help compare nursing homes. For the most current information, visit www.medicare.gov/nursinghomecompare.", methodology_style))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_quality_measures_report(self, facility_name: str, quality_data: List[Dict], benchmarks: Dict, include_comparative: bool = False) -> BytesIO:
        """Generate quality measures domain report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=self.pagesize, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.6*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        heading_style = ParagraphStyle('SectionHeading', parent=styles['Heading2'], fontSize=12, textColor=self.cms_blue, spaceAfter=10, spaceBefore=12, fontName='Helvetica-Bold', backColor=self.cms_light_gray, borderPadding=5)
        
        # CMS Header
        story.append(Paragraph("Centers for Medicare & Medicaid Services", ParagraphStyle('CMSHeader', parent=styles['Heading1'], fontSize=10, textColor=colors.black, spaceAfter=3, fontName='Helvetica-Bold')))
        story.append(Paragraph("FIVE-STAR QUALITY RATING SYSTEM - NURSING HOME COMPARE", ParagraphStyle('CMSHeader', parent=styles['Heading1'], fontSize=10, textColor=colors.black, spaceAfter=3, fontName='Helvetica-Bold')))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("QUALITY MEASURES REPORT", ParagraphStyle('ReportTitle', parent=styles['Heading1'], fontSize=16, textColor=self.cms_blue, spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')))
        story.append(Paragraph(facility_name, styles['Heading2']))
        story.append(Spacer(1, 0.15*inch))
        
        # Quality Measures
        story.append(Paragraph("QUALITY INDICATORS", heading_style))
        if quality_data:
            current = quality_data[0]
            qm_info = [
                ['Pressure Ulcer Rate', f"{current.get('pressure_ulcer_percentage', 'N/A')}%"],
                ['Urinary Tract Infections', f"{current.get('uti_percentage', 'N/A')}%"],
                ['Antipsychotic Medication Use', f"{current.get('antipsychotic_percentage', 'N/A')}%"],
                ['Hospital Readmission Rate', f"{current.get('readmission_rate', 'N/A')}%"],
                ['Hospital Transfer Rate', f"{current.get('hospital_transfer_rate', 'N/A')}%"],
            ]
            qm_table = Table(qm_info, colWidths=[3.5*inch, 1.5*inch])
            qm_table.setStyle(TableStyle([('ALIGN', (0, 0), (0, -1), 'LEFT'), ('ALIGN', (1, 0), (1, -1), 'CENTER'), ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, self.cms_light_gray])]))
            story.append(qm_table)
        
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("___" * 35, styles['Normal']))
        methodology_style = ParagraphStyle('Methodology', parent=styles['Normal'], fontSize=7, textColor=colors.grey, alignment=TA_JUSTIFY, leading=9)
        story.append(Paragraph("The Five-Star Quality Rating System provides consumers, their families, and caregivers with information to help compare nursing homes. For the most current information, visit www.medicare.gov/nursinghomecompare.", methodology_style))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_comparative_report(self, facility_name: str, facility_ratings: Dict, state_benchmark: Dict, national_benchmark: Dict) -> BytesIO:
        """Generate comparative analysis report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=self.pagesize, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.6*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        heading_style = ParagraphStyle('SectionHeading', parent=styles['Heading2'], fontSize=12, textColor=self.cms_blue, spaceAfter=10, spaceBefore=12, fontName='Helvetica-Bold', backColor=self.cms_light_gray, borderPadding=5)
        
        # CMS Header
        story.append(Paragraph("Centers for Medicare & Medicaid Services", ParagraphStyle('CMSHeader', parent=styles['Heading1'], fontSize=10, textColor=colors.black, spaceAfter=3, fontName='Helvetica-Bold')))
        story.append(Paragraph("FIVE-STAR QUALITY RATING SYSTEM - NURSING HOME COMPARE", ParagraphStyle('CMSHeader', parent=styles['Heading1'], fontSize=10, textColor=colors.black, spaceAfter=3, fontName='Helvetica-Bold')))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("COMPARATIVE ANALYSIS REPORT", ParagraphStyle('ReportTitle', parent=styles['Heading1'], fontSize=16, textColor=self.cms_blue, spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')))
        story.append(Paragraph(facility_name, styles['Heading2']))
        story.append(Spacer(1, 0.15*inch))
        
        # Comparison Table
        story.append(Paragraph("RATINGS COMPARISON", heading_style))
        comp_data = [
            ['Domain', 'Your Facility', 'State Median', 'National Median'],
            ['Overall Rating', str(facility_ratings.get('overall_rating', 'N/A')), str(state_benchmark.get('overall_median', 'N/A')), str(national_benchmark.get('overall_median', 'N/A'))],
            ['Health Inspections', str(facility_ratings.get('health_inspection_rating', 'N/A')), str(state_benchmark.get('health_inspection_median', 'N/A')), str(national_benchmark.get('health_inspection_median', 'N/A'))],
            ['Staffing', str(facility_ratings.get('staffing_rating', 'N/A')), str(state_benchmark.get('staffing_median', 'N/A')), str(national_benchmark.get('staffing_median', 'N/A'))],
            ['Quality Measures', str(facility_ratings.get('qm_rating', 'N/A')), str(state_benchmark.get('qm_median', 'N/A')), str(national_benchmark.get('qm_median', 'N/A'))],
        ]
        comp_table = Table(comp_data, colWidths=[1.8*inch, 1.3*inch, 1.3*inch, 1.3*inch])
        comp_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), self.cms_blue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, self.cms_light_gray])]))
        story.append(comp_table)
        
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("___" * 35, styles['Normal']))
        methodology_style = ParagraphStyle('Methodology', parent=styles['Normal'], fontSize=7, textColor=colors.grey, alignment=TA_JUSTIFY, leading=9)
        story.append(Paragraph("The Five-Star Quality Rating System provides consumers, their families, and caregivers with information to help compare nursing homes. For the most current information, visit www.medicare.gov/nursinghomecompare.", methodology_style))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_to_csv(self, data: List[Dict], filename: str = "report.csv") -> BytesIO:
        """Export data to CSV format"""
        output = StringIO()
        
        if not data:
            return BytesIO(b"No data to export")
        
        # Get headers from first row
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        
        writer.writeheader()
        for row in data:
            writer.writerow(row)
        
        # Convert StringIO to BytesIO
        csv_bytes = output.getvalue().encode('utf-8')
        buffer = BytesIO(csv_bytes)
        buffer.seek(0)
        return buffer
    
    def export_to_excel(self, data: List[Dict], filename: str = "report.xlsx", sheet_name: str = "Report") -> BytesIO:
        """Export data to Excel format"""
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl not available
            return self.export_to_csv(data, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet names limited to 31 chars
        
        if not data:
            ws.append(["No data to export"])
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        # Get headers from first row
        headers = list(data[0].keys())
        
        # Write header row with formatting
        ws.append(headers)
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data rows
        for row_data in data:
            row_values = [row_data.get(header, '') for header in headers]
            ws.append(row_values)
        
        # Apply borders and formatting to data rows
        thin_border = Border(
            left=Side(style='thin', color='cccccc'),
            right=Side(style='thin', color='cccccc'),
            top=Side(style='thin', color='cccccc'),
            bottom=Side(style='thin', color='cccccc')
        )
        
        for row_num, _ in enumerate(data, start=2):
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                # Alternate row colors
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_staffing_data_to_csv(self, staffing_data: List[Dict]) -> BytesIO:
        """Export staffing data to CSV format"""
        return self.export_to_csv(staffing_data, "staffing_report.csv")
    
    def export_staffing_data_to_excel(self, staffing_data: List[Dict]) -> BytesIO:
        """Export staffing data to Excel format"""
        return self.export_to_excel(staffing_data, "staffing_report.xlsx", "Staffing Data")
    
    def export_quality_data_to_csv(self, quality_data: List[Dict]) -> BytesIO:
        """Export quality measures data to CSV format"""
        return self.export_to_csv(quality_data, "quality_report.csv")
    
    def export_quality_data_to_excel(self, quality_data: List[Dict]) -> BytesIO:
        """Export quality measures data to Excel format"""
        return self.export_to_excel(quality_data, "quality_report.xlsx", "Quality Measures")
    
    def export_comparative_data_to_csv(self, comparative_data: List[Dict]) -> BytesIO:
        """Export comparative analysis data to CSV format"""
        return self.export_to_csv(comparative_data, "comparative_report.csv")
    
    def export_comparative_data_to_excel(self, comparative_data: List[Dict]) -> BytesIO:
        """Export comparative analysis data to Excel format"""
        return self.export_to_excel(comparative_data, "comparative_report.xlsx", "Comparative Analysis")
